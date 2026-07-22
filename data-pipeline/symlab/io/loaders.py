"""Loaders: raw bytes in the vault to a validated (X, y) table.

Each loader owns the quirks of ONE source, and every quirk it handles is a documented defect from
the research phase rather than a defensive guess. Three of them exist specifically because the
source lies about itself:

- The flotation file states hourly laboratory assays on 20-second rows, so a row-wise fit leaks the
  target about 13.5 times over. `load_flotation` aggregates to the hourly grid and REFUSES to return
  row-level data at all, so the leak is structurally unavailable rather than merely discouraged.
  It also parses the comma decimal separator, which a naive reader turns into wrong numbers silently.
- The water-treatment landing page states there are no missing values; the file has 591.
  `load_water_treatment` counts them and reports the count rather than imputing quietly.
- The PMLB files are Git LFS backed, so the obvious URL returns a pointer with HTTP 200.

Every loader returns a `LoadedDataset` carrying the arrays, the column semantics, the provenance and
whatever defects were encountered, so that nothing downstream has to re-derive where a number came
from.
"""
from __future__ import annotations

import gzip
import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np

from ..model.units import DIMENSIONLESS, Dims
from .sources import SOURCES, Source


@dataclass
class LoadedDataset:
    """A dataset after ingestion, with everything needed to cite and to audit it."""

    id: str
    name: str
    X: np.ndarray
    y: np.ndarray
    input_keys: list[str]
    input_display: list[str]
    input_units: list[str]
    input_dims: list[Dims]
    target_key: str
    target_display: str
    target_unit: str
    target_dims: Dims = DIMENSIONLESS
    source_id: str = ""
    citation: str = ""
    licence: str = ""
    redistribution: str = ""
    real_or_synthetic: str = "real"
    #: How many rows the SOURCE carried, before any subsample. Recorded so the app can state
    #: "4 000 of 9 568 used" rather than printing one number that contradicts the case description.
    n_rows_source: int | None = None
    defects_applied: list[str] = field(default_factory=list)
    notes: str = ""
    ground_truth_latex: str | None = None

    def __post_init__(self) -> None:
        if self.X.ndim != 2:
            raise ValueError(f"{self.id}: X must be 2-D, got shape {self.X.shape}")
        if self.y.ndim != 1:
            raise ValueError(f"{self.id}: y must be 1-D, got shape {self.y.shape}")
        if self.X.shape[0] != self.y.shape[0]:
            raise ValueError(f"{self.id}: X has {self.X.shape[0]} rows, y has {self.y.shape[0]}")
        if self.X.shape[1] != len(self.input_keys):
            raise ValueError(
                f"{self.id}: X has {self.X.shape[1]} columns but {len(self.input_keys)} keys"
            )


def _unpacked_dir(source: Source) -> Path:
    return source.path.parent / "unpacked"


def _read_zip_member(source: Source, suffixes: tuple[str, ...]) -> bytes:
    """Read the first member of an archive whose name ends in one of `suffixes`."""
    with zipfile.ZipFile(source.path) as archive:
        for name in archive.namelist():
            if name.lower().endswith(suffixes):
                return archive.read(name)
    raise FileNotFoundError(f"{source.id}: no member ending in {suffixes} inside {source.path.name}")


# --------------------------------------------------------------------------------------------
# PMLB (Feynman, Strogatz, Nikuradse)
# --------------------------------------------------------------------------------------------


def load_pmlb(dataset: str) -> LoadedDataset:
    """Load one PMLB dataset. The last column is the target, by PMLB convention."""
    source = SOURCES[f"pmlb-{dataset}"]
    with gzip.open(source.path, "rt", encoding="utf-8") as handle:
        header = handle.readline().rstrip("\n").split("\t")
        rows = np.loadtxt(handle, delimiter="\t")
    if rows.ndim == 1:
        rows = rows.reshape(1, -1)
    X, y = rows[:, :-1], rows[:, -1]
    keys = header[:-1]
    return LoadedDataset(
        id=f"pmlb-{dataset}",
        name=source.name,
        X=X, y=y,
        input_keys=keys,
        input_display=list(keys),
        input_units=["1"] * len(keys),
        input_dims=[DIMENSIONLESS] * len(keys),
        target_key=header[-1],
        target_display=header[-1],
        target_unit="1",
        source_id=source.id,
        citation=source.citation,
        licence=source.licence,
        redistribution=source.redistribution,
        notes=source.notes,
    )


# --------------------------------------------------------------------------------------------
# Flotation: the target-broadcast trap
# --------------------------------------------------------------------------------------------

HOURLY_AGGREGATION_NOTE = (
    "Aggregated to the hourly grid before any fitting. The concentrate assays are hourly laboratory "
    "measurements repeated across every 20-second process row; fitting at row level would leak the "
    "target about 13.5 times over. Row-level access is not offered by this loader."
)


def load_flotation(*, target: str = "silica") -> LoadedDataset:
    """Load the froth flotation plant data, aggregated to the hourly grid.

    The aggregation is not a modelling choice, it is a correctness requirement. Process variables are
    averaged within the hour; the assay columns are constant within the hour by construction and are
    taken as the first value, which is verified rather than assumed.
    """
    source = SOURCES["flotation-mining-process"]
    text = source.path.read_text(encoding="utf-8", errors="replace")

    # The ARFF header names the attributes; the data section follows @data.
    attributes: list[str] = []
    data_start = 0
    lines = text.splitlines()
    for index, line in enumerate(lines):
        stripped = line.strip()
        low = stripped.lower()
        if low.startswith("@attribute"):
            attributes.append(stripped.split()[1].strip("'\""))
        elif low.startswith("@data"):
            data_start = index + 1
            break

    # The decimal separator is a comma INSIDE quoted fields. Splitting on commas naively corrupts
    # every number; the file is therefore parsed field by field, honouring the quotes.
    def split_quoted(line: str) -> list[str]:
        out: list[str] = []
        current: list[str] = []
        in_quotes = False
        for char in line:
            if char in "'\"":
                in_quotes = not in_quotes
                continue
            if char == "," and not in_quotes:
                out.append("".join(current))
                current = []
                continue
            current.append(char)
        out.append("".join(current))
        return out

    stamps: list[str] = []
    numeric_rows: list[list[float]] = []
    parse_failures = 0
    for line in lines[data_start:]:
        if not line.strip():
            continue
        fields = split_quoted(line)
        if len(fields) != len(attributes):
            parse_failures += 1
            continue
        try:
            values = [float(f.replace(",", ".")) for f in fields[1:]]
        except ValueError:
            parse_failures += 1
            continue
        stamps.append(fields[0])
        numeric_rows.append(values)

    # CONTENT GUARD. The recorded source URL for this dataset was once wrong and served an unrelated
    # dataset with HTTP 200 and a plausible size. Checking the status code cannot catch that; checking
    # the schema can. If these columns are absent, the file is not the flotation dataset whatever its
    # name says.
    required = {"%_Iron_Feed", "%_Silica_Feed", "%_Silica_Concentrate", "Amina_Flow", "Ore_Pulp_pH"}
    present = set(attributes)
    if not required.issubset(present):
        raise RuntimeError(
            f"{source.id}: the file does not carry the flotation schema. Missing "
            f"{sorted(required - present)}. Found {sorted(present)[:8]}. Re-fetch from the id "
            "resolved through the OpenML API for dataset 43311."
        )
    if not numeric_rows:
        raise RuntimeError(f"{source.id}: parsed zero data rows from {len(lines)} lines.")

    matrix = np.asarray(numeric_rows, dtype=np.float64)
    columns = attributes[1:]

    # Group by hour: the timestamp string truncated to the hour.
    hour_keys = [s[:13] for s in stamps]
    order: dict[str, list[int]] = {}
    for index, key in enumerate(hour_keys):
        order.setdefault(key, []).append(index)

    aggregated = np.vstack([matrix[idx].mean(axis=0) for idx in order.values()])

    target_column = "%_Silica_Concentrate" if target == "silica" else "%_Iron_Concentrate"
    if target_column not in columns:
        matches = [c for c in columns if target in c.lower()]
        if not matches:
            raise KeyError(f"no column matching {target!r} among {columns}")
        target_column = matches[0]
    target_index = columns.index(target_column)

    # Drop BOTH concentrate assays from the inputs: predicting silica from iron is not a soft sensor,
    # it is reading the answer off the other half of the same laboratory measurement.
    excluded = {c for c in columns if "Concentrate" in c}
    keep = [i for i, c in enumerate(columns) if c not in excluded]

    return LoadedDataset(
        id="flotation-silica" if target == "silica" else "flotation-iron",
        name=source.name,
        X=aggregated[:, keep],
        y=aggregated[:, target_index],
        input_keys=[columns[i] for i in keep],
        input_display=[columns[i] for i in keep],
        input_units=["1"] * len(keep),
        input_dims=[DIMENSIONLESS] * len(keep),
        target_key=target_column,
        target_display=target_column,
        target_unit="%",
        source_id=source.id,
        citation=source.citation,
        licence=source.licence,
        redistribution=source.redistribution,
        defects_applied=[
            HOURLY_AGGREGATION_NOTE,
            f"Comma decimal separator handled explicitly; {parse_failures} malformed rows skipped.",
            "Both concentrate assay columns excluded from the inputs: predicting one from the other "
            "is reading the answer off the same laboratory measurement, not soft sensing.",
        ],
        notes=source.notes,
    )


def load_flotation_mineralogy() -> LoadedDataset:
    """The ore-mineralogy closure case: iron feed against silica feed.

    This is the case where the research measured a real discrepancy between the plant data and
    two-mineral stoichiometry. Re-derived from the rows this loader actually returns, by ordinary
    least squares on the 4097 hourly rows:

        measured      Fe = 67.0831 - 0.7363 * Si
        stoichiometry Fe = 69.94   - 0.699  * Si

    The research phase recorded 67.11 and 0.738 for the measured line. That is the same line to the
    precision the research quoted it at, and the figures above are the ones this code reproduces, so
    they are the ones stated: a number in a docstring should be recomputable from the module it sits
    in. Note the fit is weighted by the hourly broadcast; on the 309 unique (Si, Fe) pairs it is
    Fe = 65.2725 - 0.6438 * Si, which is a different question and not the one the case asks.

    Recovering the measured line and comparing it against the theoretical one is the whole point, so
    nothing here tries to reconcile them.
    """
    source = SOURCES["flotation-mining-process"]
    base = load_flotation(target="silica")
    keys = base.input_keys
    fe_idx = next(i for i, k in enumerate(keys) if "Iron_Feed" in k)
    si_idx = next(i for i, k in enumerate(keys) if "Silica_Feed" in k)
    return LoadedDataset(
        id="ore-mineralogy-closure",
        name="Ore mineralogy closure: iron feed against silica feed",
        X=base.X[:, [si_idx]],
        y=base.X[:, fe_idx],
        input_keys=[keys[si_idx]],
        input_display=["\\%\\,\\mathrm{SiO_2}\\ \\mathrm{feed}"],
        input_units=["%"],
        input_dims=[DIMENSIONLESS],
        target_key=keys[fe_idx],
        target_display="\\%\\,\\mathrm{Fe}\\ \\mathrm{feed}",
        target_unit="%",
        source_id=source.id,
        citation=source.citation,
        licence=source.licence,
        redistribution=source.redistribution,
        ground_truth_latex=r"\%\mathrm{Fe} = 69.94 - 0.699\,\%\mathrm{SiO_2} \quad\text{(two-mineral stoichiometry)}",
        defects_applied=base.defects_applied,
        notes=(
            "The stoichiometric prediction assumes the ore is only hematite and quartz. The measured "
            "line differs, and the size of that difference is a statement about the other minerals "
            "present. The lab reports both lines rather than choosing one."
        ),
    )


# --------------------------------------------------------------------------------------------
# UCI sources
# --------------------------------------------------------------------------------------------


def _read_xlsx_sheet(data: bytes) -> tuple[list[str], np.ndarray]:
    """Read the first sheet of an xlsx into a header and a float matrix."""
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(v).strip() for v in next(rows)]
    values = [[float(v) for v in row] for row in rows if row and row[0] is not None]
    workbook.close()
    return header, np.asarray(values, dtype=np.float64)


def _read_xls_sheet(data: bytes) -> tuple[list[str], np.ndarray]:
    """Read the first sheet of a legacy .xls into a header and a float matrix."""
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    header = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    values = [
        [float(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        for r in range(1, sheet.nrows)
        if sheet.cell_value(r, 0) != ""
    ]
    return header, np.asarray(values, dtype=np.float64)


def load_ccpp() -> LoadedDataset:
    """Combined cycle power plant. Inputs AT, V, AP, RH; target PE.

    The UCI archive ships the data as a spreadsheet with five folds of the SAME 9,568 records in
    different orders (the `Folds5x2_pp` name says so). Only the first sheet is read; concatenating
    the folds would silently five-fold-duplicate every row and make any held-out split meaningless.
    """
    source = SOURCES["ccpp"]
    with zipfile.ZipFile(source.path) as archive:
        name = next(n for n in archive.namelist() if n.lower().endswith(".xlsx"))
        header, rows = _read_xlsx_sheet(archive.read(name))
    return LoadedDataset(
        id="ccpp-derating", name=source.name,
        X=rows[:, :4], y=rows[:, 4],
        input_keys=header[:4],
        input_display=["T_{amb}", "V_{exh}", "p_{amb}", r"\mathrm{RH}"],
        input_units=["degC", "cmHg", "mbar", "%"],
        input_dims=[DIMENSIONLESS] * 4,
        target_key=header[4], target_display="P_e", target_unit="MW",
        source_id=source.id, citation=source.citation, licence=source.licence,
        redistribution=source.redistribution,
        defects_applied=[
            "Only the first of the five shuffled folds in the workbook is read. The folds are "
            "reorderings of the same 9,568 records; concatenating them would duplicate every row "
            "five times and destroy any held-out split.",
        ],
        notes=source.notes,
    )


def load_concrete() -> LoadedDataset:
    """Concrete compressive strength. Eight mix and age inputs, one strength target."""
    source = SOURCES["concrete"]
    with zipfile.ZipFile(source.path) as archive:
        name = next(n for n in archive.namelist() if n.lower().endswith((".xls", ".xlsx")))
        payload = archive.read(name)
        header, rows = (_read_xlsx_sheet if name.lower().endswith(".xlsx") else _read_xls_sheet)(payload)
    short = [
        "cement", "slag", "fly_ash", "water", "superplasticizer",
        "coarse_aggregate", "fine_aggregate", "age",
    ]
    return LoadedDataset(
        id="concrete-abrams", name=source.name,
        X=rows[:, :-1], y=rows[:, -1],
        input_keys=short,
        input_display=[
            "c", "s", "f", "w", "sp", "g_{coarse}", "g_{fine}", "t_{age}",
        ],
        input_units=["kg/m^3"] * 7 + ["d"],
        input_dims=[DIMENSIONLESS] * 8,
        target_key="strength", target_display="f_c", target_unit="MPa",
        source_id=source.id, citation=source.citation, licence=source.licence,
        redistribution=source.redistribution,
        ground_truth_latex=(
            r"f_c \approx \frac{A}{B^{w/c}} \quad \text{(Abrams), with a }"
            r" \ln(t_{age}) \text{ maturity term}"
        ),
        defects_applied=[
            "The workbook column names carry units inside the label; they are replaced by short "
            "machine keys and the units are recorded separately, so a discovered expression renders "
            "as mathematics rather than as a spreadsheet caption.",
        ],
        notes=source.notes,
    )


def load_water_treatment() -> LoadedDataset:
    """Urban wastewater treatment. Carries the exact removal-efficiency identity.

    The landing page claims no missing values. The file has 591, marked with `?`. They are counted,
    reported and the affected rows dropped, because imputing into an EXACT identity would manufacture
    the very relationship the case exists to recover.
    """
    source = SOURCES["water-treatment"]
    raw = _read_zip_member(source, (".data", ".csv", ".txt"))
    text = raw.decode("utf-8", errors="replace")
    lines = [ln for ln in text.splitlines() if ln.strip()]
    missing = sum(line.count("?") for line in lines)

    # The UCI file has no header row; the attribute order is fixed by the documentation.
    columns = [
        "Q-E", "ZN-E", "PH-E", "DBO-E", "DQO-E", "SS-E", "SSV-E", "SED-E", "COND-E",
        "PH-P", "DBO-P", "SS-P", "SSV-P", "SED-P", "COND-P",
        "PH-D", "DBO-D", "DQO-D", "SS-D", "SSV-D", "SED-D", "COND-D",
        "PH-S", "DBO-S", "DQO-S", "SS-S", "SSV-S", "SED-S", "COND-S",
        "RD-DBO-P", "RD-SS-P", "RD-SED-P", "RD-DBO-S", "RD-DQO-S",
        "RD-DBO-G", "RD-DQO-G", "RD-SS-G", "RD-SED-G",
    ]
    kept: list[list[float]] = []
    for line in lines:
        fields = line.split(",")[1:]  # first field is the date label
        if len(fields) != len(columns) or "?" in fields:
            continue
        try:
            kept.append([float(v) for v in fields])
        except ValueError:
            continue
    matrix = np.asarray(kept, dtype=np.float64)

    # The identity: global BOD removal from the inlet and outlet BOD.
    inlet = columns.index("DBO-E")
    outlet = columns.index("DBO-S")
    target = columns.index("RD-DBO-G")
    return LoadedDataset(
        id="wwtp-removal-identity",
        name=source.name,
        X=matrix[:, [inlet, outlet]],
        y=matrix[:, target],
        input_keys=["DBO-E", "DBO-S"],
        input_display=[r"\mathrm{BOD}_{in}", r"\mathrm{BOD}_{out}"],
        input_units=["mg/L", "mg/L"],
        input_dims=[DIMENSIONLESS, DIMENSIONLESS],
        target_key="RD-DBO-G", target_display=r"\mathrm{RD}_{\mathrm{BOD}}", target_unit="%",
        source_id=source.id, citation=source.citation, licence=source.licence,
        redistribution=source.redistribution,
        ground_truth_latex=r"\mathrm{RD} = 100\,\frac{\mathrm{BOD}_{in} - \mathrm{BOD}_{out}}{\mathrm{BOD}_{in}}",
        defects_applied=[
            f"The UCI landing page states there are no missing values. The shipped file contains "
            f"{missing} occurrences of '?'. Rows containing any are DROPPED, not imputed: imputing "
            f"into an exact identity would manufacture the relationship this case exists to recover. "
            f"{len(kept)} of {len(lines)} rows retained.",
        ],
        notes=source.notes,
    )


def load_gas_turbine() -> LoadedDataset:
    """Gas turbine NOx emissions, concatenated across the published years."""
    source = SOURCES["gas-turbine-emissions"]
    frames: list[np.ndarray] = []
    header: list[str] = []
    with zipfile.ZipFile(source.path) as archive:
        for name in sorted(n for n in archive.namelist() if n.lower().endswith(".csv")):
            payload = archive.read(name).decode("utf-8", errors="replace")
            lines = [ln for ln in payload.splitlines() if ln.strip()]
            if not header:
                header = [h.strip() for h in lines[0].split(",")]
            frames.append(np.array([[float(v) for v in ln.split(",")] for ln in lines[1:]],
                                   dtype=np.float64))
    matrix = np.vstack(frames)
    nox = header.index("NOX")
    # Carbon monoxide is dropped: it is a co-measured emission, not a driver of NOx, and leaving it
    # in lets a search "explain" NOx with another pollutant instead of with the combustion physics.
    co = header.index("CO")
    keep = [i for i in range(len(header)) if i not in (nox, co)]
    return LoadedDataset(
        id="gasturbine-nox", name=source.name,
        X=matrix[:, keep], y=matrix[:, nox],
        input_keys=[header[i] for i in keep],
        input_display=[header[i] for i in keep],
        input_units=["1"] * len(keep),
        input_dims=[DIMENSIONLESS] * len(keep),
        target_key="NOX", target_display=r"\mathrm{NO}_x", target_unit="mg/m^3",
        source_id=source.id, citation=source.citation, licence=source.licence,
        redistribution=source.redistribution,
        defects_applied=[
            "Carbon monoxide excluded from the inputs: it is a co-measured emission rather than a "
            "driver, and including it lets a search explain one pollutant with another.",
        ],
        notes=source.notes,
    )


def load_nikuradse() -> LoadedDataset:
    """Nikuradse rough-pipe friction, 362 measured points, via the PMLB carrier."""
    base = load_pmlb("nikuradse_1")
    base.id = "nikuradse-friction"
    base.name = "Nikuradse rough-pipe friction (362 measured points)"
    base.input_display = [r"r/k", r"\log_{10}\mathrm{Re}"]
    base.target_display = r"\lambda"
    base.notes = (
        "The real-data twin of the friction-factor generator. The transitional hump in these "
        "measurements is genuinely not reproduced by the Colebrook correlation, which makes this the "
        "case where a correlation everyone uses is visibly incomplete. The Princeton host cited in "
        "the literature returns HTTP 404; PMLB carries the same points under MIT."
    )
    return base


#: Loaders by case id, so the registry can resolve a case to its data without a chain of imports.
LOADERS = {
    "flotation-silica": lambda: load_flotation(target="silica"),
    "ore-mineralogy-closure": load_flotation_mineralogy,
    "ccpp-derating": load_ccpp,
    "concrete-abrams": load_concrete,
    "wwtp-removal-identity": load_water_treatment,
    "gasturbine-nox": load_gas_turbine,
    "nikuradse-friction": load_nikuradse,
}
